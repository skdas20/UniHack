"""The 252-column delivery schema.

Nothing here is hardcoded. The schema is *read at runtime* from the Expected
Output sheet the organisers ship. If the evaluation dataset arrives with a
different, longer, or reordered header list, this module adapts to it and the
rest of the pipeline follows -- which is the whole point of the submission
note that a prototype "should not be mocked, hardcoded, or built as a basic
simulation".

The only structural assumption is the *naming convention* of the repeating
column families (``ATTRIBUTE_LABEL 1`` / ``ATTRIBUTE_VALUE 1`` /
``ATTRIBUTE_UOM 1``, ``ITEM_FEATURES_1``, ``Ref URL 1``, ``Alternate Image 1``),
and even that is discovered by regex rather than assumed to be 50/20/5/4 wide.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from functools import cached_property
from pathlib import Path

# Repeating column families, discovered by pattern rather than by count.
_FAMILIES: dict[str, re.Pattern[str]] = {
    "attribute_label": re.compile(r"^ATTRIBUTE_LABEL (\d+)$"),
    "attribute_value": re.compile(r"^ATTRIBUTE_VALUE (\d+)$"),
    "attribute_uom": re.compile(r"^ATTRIBUTE_UOM (\d+)$"),
    "item_feature": re.compile(r"^ITEM_FEATURES_(\d+)$"),
    "ref_url": re.compile(r"^Ref URL (\d+)$"),
    "alternate_image": re.compile(r"^Alternate Image (\d+)$"),
}

# Columns that are a verbatim passthrough of the raw input row. Copying these
# is not enrichment, but getting them wrong invalidates the join back to the
# distributor's system, so they are treated as first-class.
PASSTHROUGH = (
    "Mfg_Part_Num",
    "Part_Desc",
    "E1_Brand",
    "Unilog_Brand",
    "DIB_Brand",
    "Part_Manuf",
)

# Columns that are a function of no input column whatsoever. See
# docs/DERIVED_RULES.md section 7. We refuse to invent these.
NOT_DERIVABLE: dict[str, str] = {
    "PART_NUMBER": "distributor ERP item master",
    "SKU - MY_PART_NUMBER": "distributor ERP SKU master",
    "List Price": "distributor pricing system / manufacturer price file",
    "UPC": "GS1 registry or manufacturer packaging data",
    "EAN": "GS1 registry or manufacturer packaging data",
    "GTIN": "GS1 registry or manufacturer packaging data",
    "Country Of Origin": "manufacturer compliance documentation",
    "Selling Qty": "distributor ERP unit-of-sale configuration",
    "Prop 65": "manufacturer compliance documentation",
    "MTR": "manufacturer mill test report",
}


@dataclass(frozen=True)
class OutputSchema:
    """The delivery-format header list, plus structure discovered within it."""

    headers: tuple[str, ...]
    source_path: Path | None = None

    # ---------- construction ----------

    @classmethod
    def from_csv(cls, path: str | Path) -> "OutputSchema":
        path = Path(path)
        with path.open(newline="", encoding="utf-8-sig") as fh:
            headers = next(csv.reader(fh))
        return cls(headers=tuple(h.strip() for h in headers), source_path=path)

    # ---------- repeating families ----------

    @cached_property
    def _family_index(self) -> dict[str, dict[int, int]]:
        """family name -> {slot number: column position}."""
        out: dict[str, dict[int, int]] = {k: {} for k in _FAMILIES}
        for pos, header in enumerate(self.headers):
            for fam, pat in _FAMILIES.items():
                m = pat.match(header)
                if m:
                    out[fam][int(m.group(1))] = pos
                    break
        return out

    def slots(self, family: str) -> list[int]:
        """Sorted slot numbers available for a repeating family."""
        return sorted(self._family_index[family])

    @cached_property
    def n_attribute_slots(self) -> int:
        """How many attribute triples the sheet actually provides."""
        labels = set(self._family_index["attribute_label"])
        values = set(self._family_index["attribute_value"])
        uoms = set(self._family_index["attribute_uom"])
        return len(labels & values & uoms)

    @cached_property
    def n_feature_slots(self) -> int:
        return len(self._family_index["item_feature"])

    @cached_property
    def n_ref_url_slots(self) -> int:
        return len(self._family_index["ref_url"])

    @cached_property
    def n_alternate_image_slots(self) -> int:
        return len(self._family_index["alternate_image"])

    def attribute_headers(self, slot: int) -> tuple[str, str, str]:
        """(label_header, value_header, uom_header) for a 1-based slot."""
        return (
            f"ATTRIBUTE_LABEL {slot}",
            f"ATTRIBUTE_VALUE {slot}",
            f"ATTRIBUTE_UOM {slot}",
        )

    # ---------- lookups ----------

    @cached_property
    def position(self) -> dict[str, int]:
        return {h: i for i, h in enumerate(self.headers)}

    def __contains__(self, header: object) -> bool:
        return header in self.position

    def __len__(self) -> int:
        return len(self.headers)

    def __iter__(self):
        return iter(self.headers)

    def blank_record(self) -> dict[str, str]:
        """An empty record with every header present, in order."""
        return {h: "" for h in self.headers}

    # ---------- validation ----------

    def validate(self, record: dict[str, str]) -> list[str]:
        """Return a list of structural problems with a produced record.

        An empty list means the record is safe to write into the delivery sheet.
        """
        problems: list[str] = []
        missing = [h for h in self.headers if h not in record]
        if missing:
            problems.append(
                f"{len(missing)} header(s) absent from record, first few: {missing[:5]}"
            )
        extra = [k for k in record if k not in self.position]
        if extra:
            problems.append(
                f"{len(extra)} unknown key(s) not in schema, first few: {extra[:5]}"
            )
        # An attribute value or UOM with no label is unreadable downstream.
        for slot in self.slots("attribute_value"):
            lab, val, uom = self.attribute_headers(slot)
            if lab not in record:
                continue
            if not record.get(lab, "").strip():
                if record.get(val, "").strip() or record.get(uom, "").strip():
                    problems.append(f"slot {slot}: value/UOM present with no label")
        return problems


@dataclass
class RawRow:
    """One row of the 6-column input, plus its position in the file."""

    index: int
    data: dict[str, str] = field(default_factory=dict)

    def get(self, key: str, default: str = "") -> str:
        return (self.data.get(key) or default).strip()

    @property
    def mpn(self) -> str:
        return self.get("Mfg_Part_Num")

    @property
    def desc(self) -> str:
        return self.get("Part_Desc")


def load_input(path: str | Path) -> list[RawRow]:
    """Read the working dataset. Tolerates CSV or XLSX, any column order."""
    path = Path(path)
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        out = []
        for i, values in enumerate(rows):
            if values is None or all(v is None for v in values):
                continue
            data = {
                h: ("" if v is None else str(v))
                for h, v in zip(header, values)
                if h
            }
            out.append(RawRow(index=i, data=data))
        wb.close()
        return out

    with path.open(newline="", encoding="utf-8-sig") as fh:
        return [
            RawRow(index=i, data={k.strip(): (v or "") for k, v in rec.items() if k})
            for i, rec in enumerate(csv.DictReader(fh))
        ]


def load_gold(path: str | Path) -> list[dict[str, str]]:
    """Read the worked example rows out of the Expected Output sheet.

    These are the only labelled ground truth in the pack (the 200-row
    Input-vs-Output workbook referenced by the Solution Guide is not published
    on the portal), so they are treated as the format specification.
    """
    path = Path(path)
    with path.open(newline="", encoding="utf-8-sig") as fh:
        return [
            {k.strip(): (v or "").strip() for k, v in rec.items() if k}
            for rec in csv.DictReader(fh)
        ]
